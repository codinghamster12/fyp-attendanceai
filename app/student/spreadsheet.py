from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
import os
import sqlite3

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#database connection
conn = sqlite3.connect('../db.sqlite3')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb['attendance']
    # sheet[ord() + '1']
    for col_index in range(1, 200):
    	col = get_column_letter(col_index)
        
    	if sheet['%s%s' % (col,1)].value is None:
    		col2 = get_column_letter(col_index - 1)
    		# print sheet.cell('%s%s'% (col2, 1)).value
    		if sheet['%s%s' % (col2,1)].value != currentDate:
    			sheet['%s%s' % (col,1)] = currentDate
    		break

    #saving the file
    wb.save(filename = "reports.xlsx")

else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM student_student ORDER BY Registration_No ASC")

    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "attendance"
    ws1.append(('Registration_No', 'Name', currentDate))
    ws1.append(('', '', ''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)
